import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('CSE Innovation Day Project Details -2026.xlsx')
ws = wb['CSE']

# Find the header row by looking for common column names
header_row = None
for row_idx in range(1, 10):
    row_values = [cell.value for cell in ws[row_idx]]
    row_str = ' '.join([str(v) for v in row_values if v])
    if 'PROJECT' in row_str.upper() or 'TITLE' in row_str.upper():
        header_row = row_idx
        break

if not header_row:
    header_row = 1

# Get headers
headers = [cell.value for cell in ws[header_row]]
print(f"Row {header_row} Headers:")
for idx, h in enumerate(headers):
    if h:
        print(f"  Col {idx}: {h}")

# Find project title column
title_col = None
desc_col = None
for idx, h in enumerate(headers):
    if h and ('PROJECT' in str(h).upper() or 'TITLE' in str(h).upper()):
        title_col = idx
    if h and ('DESC' in str(h).upper() or 'ABSTRACT' in str(h).upper()):
        desc_col = idx

print(f"\nTitle Column: {title_col}, Description Column: {desc_col}\n")

# Extract projects
projects = []
for row in ws.iter_rows(min_row=header_row+1, max_row=header_row+50, values_only=True):
    if title_col is not None and row[title_col]:
        title = str(row[title_col]).strip()
        desc = str(row[desc_col]).strip() if desc_col and row[desc_col] else 'N/A'
        if title and len(title) > 3:
            projects.append({'title': title, 'description': desc})
            print(f"\n{len(projects)}. {title}")
            if desc != 'N/A':
                print(f"   Description: {desc[:200]}...")

print(f"\n\nTotal unique projects: {len(projects)}")
